"""
生成 BeadForge 功能清单 Excel（包含已完成 / 部分完成 / 未完成）
输出: docs/feature-status.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).resolve().parent.parent / "docs" / "feature-status.xlsx"
OUT.parent.mkdir(parents=True, exist_ok=True)

# ---------- 样式 ----------
HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
DONE_FILL   = PatternFill("solid", fgColor="D4EFDF")  # 绿
PART_FILL   = PatternFill("solid", fgColor="FCF3CF")  # 黄
TODO_FILL   = PatternFill("solid", fgColor="FADBD8")  # 红
CENTER      = Alignment(horizontal="center", vertical="center")
LEFT_WRAP   = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN        = Side(border_style="thin", color="BDC3C7")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DONE = "✅ 已完成"
PART = "⚠️ 部分完成"
TODO = "❌ 未完成"

def status_fill(s):
    return DONE_FILL if s == DONE else (PART_FILL if s == PART else TODO_FILL)

def write_header(ws, headers, widths):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = widths[i-1]
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"

def write_row(ws, row_idx, values, status_col_idx):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row_idx, column=i, value=v)
        c.border = BORDER
        c.alignment = LEFT_WRAP if i != status_col_idx else CENTER
        if i == status_col_idx:
            c.fill = status_fill(v)
            c.font = Font(bold=True)

# ============================================================
# Sheet 1: 模块总览
# ============================================================
wb = Workbook()
ws1 = wb.active
ws1.title = "模块总览"
write_header(ws1, ["#", "模块", "覆盖", "接口数", "状态", "备注"], [4, 18, 28, 8, 12, 50])

modules = [
    ("认证",       "注册 / 登录",                              2,  DONE, "JWT 24h；公开接口"),
    ("用户",       "资料 / 统计 / 改密 / 社区档案",             5,  DONE, "邮箱字段无验证"),
    ("收货地址",   "CRUD + 默认地址",                          6,  DONE, "默认地址唯一性事务保证"),
    ("作品",       "创建 / 编辑 / 删除 / 复制 / 公开列表",      8,  DONE, "已去 mock"),
    ("图纸市场",   "浏览 / 发布 / 购买 / 已购列表",             4,  DONE, "免费 + 付费两条路径"),
    ("评论",       "列表 / 发布 / 删除（含回复）",              3,  DONE, "已去 mock"),
    ("点赞",       "点赞 / 取消 / 给出 / 收到 / 公开",         6,  DONE, "幂等；同步 like_count"),
    ("收藏",       "添加 / 取消 / 检查 / 列表",                4,  DONE, "已去 mock"),
    ("AI 生图",    "豆包 Seedream → 像素化",                   1,  DONE, "24 色拼豆调色板匹配"),
    ("文件上传",   "图片 / 视频上传",                          1,  DONE, "本地 ./uploads，最大 20MB"),
    ("动态",       "推荐 / 最新 / 关注 / 我的 / 他人",          5,  DONE, "已去 mock"),
    ("关注关系",   "关注 / 取关 / 粉丝 / 关注列表",            5,  DONE, ""),
    ("通知",       "列表 / 未读 / 已读 / 全已读 / 删除",        5,  DONE, ""),
    ("材料商品",   "列表 / 详情",                              2,  DONE, "公开接口"),
    ("订单",       "创建 / 支付 / 发货 / 收货 / 取消 / 退款",   9,  PART, "支付为模拟；退款仅状态变更"),
    ("充值支付",   "创建订单 / 确认 / 配置查询",                3,  PART, "微信/支付宝下单 + 回调待接"),
    ("钱包",       "余额 / 签到 / 充值 / 买图纸 / 流水",        5,  DONE, "拼豆币体系"),
    ("管理后台",   "统计 / 各模块 CRUD / API 配置",            15, DONE, "图纸/动态仅支持删除"),
    ("发现页配置", "Banner / Tab / 全局文案",                  11, DONE, "新加 admin 可在线管理"),
    ("字典管理",   "字典项 CRUD + reload",                     5,  DONE, "订单/工单/钱包等枚举"),
    ("反馈工单",   "列表 / 详情 / 创建 / 用户回复",             4,  PART, "缺管理后台回复入口"),
    ("私信 (DM)",  "前端 DirectMessageScreen",                 0,  TODO, "无后端模块；待设计"),
    ("邮箱验证",   "注册时邮箱验证码",                          0,  TODO, "register 不校验邮箱"),
]

total = 0
for i, (mod, scope, n, status, note) in enumerate(modules, 2):
    write_row(ws1, i, [i-1, mod, scope, n, status, note], 5)
    total += n

ws1.cell(row=len(modules)+2, column=2, value="合计").font = Font(bold=True)
ws1.cell(row=len(modules)+2, column=4, value=total).font = Font(bold=True)
for col in range(1, 7):
    ws1.cell(row=len(modules)+2, column=col).fill = PatternFill("solid", fgColor="ECF0F1")
    ws1.cell(row=len(modules)+2, column=col).border = BORDER

# ============================================================
# Sheet 2: 后端接口清单
# ============================================================
ws2 = wb.create_sheet("后端接口")
write_header(ws2, ["#", "模块", "方法", "路径", "说明", "认证", "状态", "备注"],
             [4, 14, 8, 36, 36, 8, 12, 30])

endpoints = [
    # 认证
    ("认证", "POST",   "/auth/register",                          "注册",                              "公开", DONE, ""),
    ("认证", "POST",   "/auth/login",                             "登录",                              "公开", DONE, ""),
    # 用户
    ("用户", "GET",    "/user/profile",                           "我的资料",                          "需登录", DONE, ""),
    ("用户", "PUT",    "/user/profile",                           "改资料",                            "需登录", DONE, ""),
    ("用户", "GET",    "/user/stats",                             "我的统计",                          "需登录", DONE, ""),
    ("用户", "POST",   "/user/change-password",                   "修改密码",                          "需登录", DONE, ""),
    ("用户", "GET",    "/user/community/{name}",                  "社区用户档案",                       "公开", DONE, ""),
    # 地址
    ("收货地址", "GET",    "/addresses",                          "地址列表",                          "需登录", DONE, ""),
    ("收货地址", "GET",    "/addresses/{id}",                     "地址详情",                          "需登录", DONE, ""),
    ("收货地址", "POST",   "/addresses",                          "新增地址",                          "需登录", DONE, ""),
    ("收货地址", "PUT",    "/addresses/{id}",                     "编辑地址",                          "需登录", DONE, ""),
    ("收货地址", "DELETE", "/addresses/{id}",                     "删除地址",                          "需登录", DONE, ""),
    ("收货地址", "POST",   "/addresses/{id}/default",             "设为默认",                          "需登录", DONE, ""),
    # 作品
    ("作品", "POST",   "/designs",                                "创建作品",                          "需登录", DONE, ""),
    ("作品", "GET",    "/designs/public/list",                    "公开作品列表",                      "公开", DONE, ""),
    ("作品", "GET",    "/designs/public/{id}",                    "作品详情",                          "公开", DONE, ""),
    ("作品", "GET",    "/designs/public/by-user/{userId}",        "某用户公开作品",                    "公开", DONE, ""),
    ("作品", "GET",    "/designs/my",                             "我的作品",                          "需登录", DONE, ""),
    ("作品", "PUT",    "/designs/{id}",                           "编辑作品",                          "需登录", DONE, ""),
    ("作品", "DELETE", "/designs/{id}",                           "删除作品",                          "需登录", DONE, ""),
    ("作品", "POST",   "/designs/{id}/duplicate",                 "复制作品",                          "需登录", DONE, ""),
    # 图纸市场
    ("图纸市场", "GET",    "/patterns/list",                      "图纸市场列表",                      "公开", DONE, ""),
    ("图纸市场", "POST",   "/patterns/publish",                   "发布图纸",                          "需登录", DONE, ""),
    ("图纸市场", "POST",   "/patterns/{id}/buy",                  "购买/下载图纸",                     "需登录", DONE, ""),
    ("图纸市场", "GET",    "/patterns/purchased",                 "我的已购图纸",                      "需登录", DONE, ""),
    # 评论
    ("评论", "GET",    "/comments",                               "评论列表",                          "公开", DONE, ""),
    ("评论", "POST",   "/comments",                               "发表评论",                          "需登录", DONE, ""),
    ("评论", "DELETE", "/comments/{id}",                          "删除评论",                          "需登录", DONE, ""),
    # 点赞
    ("点赞", "POST",   "/likes/{type}/{id}",                      "点赞",                              "需登录", DONE, ""),
    ("点赞", "DELETE", "/likes/{type}/{id}",                      "取消点赞",                          "需登录", DONE, ""),
    ("点赞", "GET",    "/likes/check/{type}/{id}",                "是否已赞",                          "需登录", DONE, ""),
    ("点赞", "GET",    "/likes/given",                            "我给的赞",                          "需登录", DONE, ""),
    ("点赞", "GET",    "/likes/by-user/{userId}",                 "某用户给的赞",                      "公开", DONE, ""),
    ("点赞", "GET",    "/likes/received",                         "我收到的赞",                        "需登录", DONE, ""),
    # 收藏
    ("收藏", "POST",   "/favorites/{type}/{id}",                  "添加收藏",                          "需登录", DONE, ""),
    ("收藏", "DELETE", "/favorites/{type}/{id}",                  "取消收藏",                          "需登录", DONE, ""),
    ("收藏", "GET",    "/favorites/check/{type}/{id}",            "是否已收藏",                        "需登录", DONE, ""),
    ("收藏", "GET",    "/favorites",                              "我的收藏列表",                      "需登录", DONE, ""),
    # AI
    ("AI 生图", "POST", "/ai/generate-image",                     "豆包文生图 → 拼豆 grid",            "需登录", DONE, "调用豆包 Ark API"),
    # 上传
    ("文件上传", "POST", "/upload/image",                         "上传图片/视频",                     "需登录", DONE, "本地存储"),
    # 动态
    ("动态", "GET",    "/feeds/list",                             "动态列表",                          "公开", DONE, ""),
    ("动态", "GET",    "/feeds/mine",                             "我发布的动态",                      "需登录", DONE, ""),
    ("动态", "GET",    "/feeds/by-user/{userId}",                 "某用户的动态",                      "公开", DONE, ""),
    ("动态", "GET",    "/feeds/following",                        "关注的人的动态",                    "需登录", DONE, ""),
    ("动态", "POST",   "/feeds",                                  "发布动态",                          "需登录", DONE, ""),
    # 关注
    ("关注关系", "POST",   "/follow/{targetUserId}",              "关注",                              "需登录", DONE, ""),
    ("关注关系", "DELETE", "/follow/{targetUserId}",              "取消关注",                          "需登录", DONE, ""),
    ("关注关系", "GET",    "/follow/check/{targetUserId}",        "是否已关注",                        "需登录", DONE, ""),
    ("关注关系", "GET",    "/follow/followers",                   "我的粉丝",                          "需登录", DONE, ""),
    ("关注关系", "GET",    "/follow/following",                   "我关注的人",                        "需登录", DONE, ""),
    # 通知
    ("通知", "GET",    "/notifications",                          "通知列表",                          "需登录", DONE, ""),
    ("通知", "GET",    "/notifications/unread-count",             "未读数",                            "需登录", DONE, ""),
    ("通知", "POST",   "/notifications/{id}/read",                "标记已读",                          "需登录", DONE, ""),
    ("通知", "POST",   "/notifications/read-all",                 "全部已读",                          "需登录", DONE, ""),
    ("通知", "DELETE", "/notifications/{id}",                     "删除通知",                          "需登录", DONE, ""),
    # 商品
    ("材料商品", "GET",    "/products/list",                      "商品列表",                          "公开", DONE, ""),
    ("材料商品", "GET",    "/products/{id}",                      "商品详情",                          "公开", DONE, ""),
    # 订单
    ("订单", "POST",   "/orders",                                 "创建订单",                          "需登录", DONE, ""),
    ("订单", "GET",    "/orders",                                 "订单列表",                          "需登录", DONE, ""),
    ("订单", "GET",    "/orders/stat-counts",                     "各状态订单数",                      "需登录", DONE, ""),
    ("订单", "GET",    "/orders/{id}",                            "订单详情",                          "需登录", DONE, ""),
    ("订单", "POST",   "/orders/{id}/pay",                        "支付订单",                          "需登录", PART, "模拟支付，无真实接入"),
    ("订单", "POST",   "/orders/{id}/ship",                       "发货",                              "需登录", DONE, ""),
    ("订单", "POST",   "/orders/{id}/receive",                    "确认收货",                          "需登录", DONE, ""),
    ("订单", "POST",   "/orders/{id}/cancel",                     "取消订单",                          "需登录", DONE, ""),
    ("订单", "POST",   "/orders/{id}/refund",                     "申请售后",                          "需登录", PART, "仅改状态，无真实退款"),
    # 充值支付
    ("充值支付", "POST",   "/payment/create-order",               "创建充值订单",                      "需登录", PART, "微信/支付宝下单 API 是 TODO"),
    ("充值支付", "POST",   "/payment/confirm",                    "确认支付",                          "需登录", PART, "回调验签未实现"),
    ("充值支付", "GET",    "/payment/config",                     "支付配置查询",                      "需登录", DONE, ""),
    # 钱包
    ("钱包", "GET",    "/wallet/balance",                         "钱包余额",                          "需登录", DONE, ""),
    ("钱包", "POST",   "/wallet/sign-in",                         "每日签到",                          "需登录", DONE, ""),
    ("钱包", "POST",   "/wallet/charge",                          "充值（模拟）",                      "需登录", PART, "纯模拟，正式走 /payment"),
    ("钱包", "POST",   "/wallet/buy-pattern/{id}",                "拼豆币买图纸",                      "需登录", DONE, ""),
    ("钱包", "GET",    "/wallet/logs",                            "流水",                              "需登录", DONE, ""),
    # 管理
    ("管理后台", "GET",    "/admin/stats",                        "全局统计",                          "ADMIN", DONE, ""),
    ("管理后台", "GET",    "/admin/users",                        "用户列表",                          "ADMIN", DONE, ""),
    ("管理后台", "DELETE", "/admin/users/{id}",                   "删除用户",                          "ADMIN", DONE, ""),
    ("管理后台", "GET",    "/admin/designs",                      "作品列表",                          "ADMIN", DONE, ""),
    ("管理后台", "DELETE", "/admin/designs/{id}",                 "删除作品",                          "ADMIN", DONE, ""),
    ("管理后台", "POST",   "/admin/products",                     "新增商品",                          "ADMIN", DONE, ""),
    ("管理后台", "PUT",    "/admin/products/{id}",                "更新商品",                          "ADMIN", DONE, ""),
    ("管理后台", "DELETE", "/admin/products/{id}",                "删除商品",                          "ADMIN", DONE, ""),
    ("管理后台", "DELETE", "/admin/patterns/{id}",                "下架图纸",                          "ADMIN", PART, "缺新增/编辑/上下架"),
    ("管理后台", "DELETE", "/admin/feeds/{id}",                   "删除动态",                          "ADMIN", PART, "缺编辑/置顶/精选"),
    ("管理后台", "GET",    "/admin/api-config",                   "API 配置列表（脱敏）",              "ADMIN", DONE, ""),
    ("管理后台", "GET",    "/admin/api-config/{id}/reveal",       "查看完整密钥",                      "ADMIN", DONE, "写审计日志"),
    ("管理后台", "POST",   "/admin/api-config",                   "新增 API 配置",                     "ADMIN", DONE, ""),
    ("管理后台", "PUT",    "/admin/api-config/{id}",              "更新 API 配置",                     "ADMIN", DONE, ""),
    ("管理后台", "DELETE", "/admin/api-config/{id}",              "删除 API 配置",                     "ADMIN", DONE, ""),
    # 发现页
    ("发现页配置", "GET",    "/discovery/home",                   "首屏 payload",                      "公开", DONE, ""),
    ("发现页配置", "GET",    "/discovery/banners",                "Banner 列表",                       "公开", DONE, ""),
    ("发现页配置", "POST",   "/discovery/banners",                "新建 Banner",                       "ADMIN", DONE, ""),
    ("发现页配置", "PUT",    "/discovery/banners/{id}",           "更新 Banner",                       "ADMIN", DONE, ""),
    ("发现页配置", "DELETE", "/discovery/banners/{id}",           "删除 Banner",                       "ADMIN", DONE, ""),
    ("发现页配置", "GET",    "/discovery/tabs",                   "Tab 列表",                          "公开", DONE, ""),
    ("发现页配置", "POST",   "/discovery/tabs",                   "新建 Tab",                          "ADMIN", DONE, ""),
    ("发现页配置", "PUT",    "/discovery/tabs/{id}",              "更新 Tab",                          "ADMIN", DONE, ""),
    ("发现页配置", "DELETE", "/discovery/tabs/{id}",              "删除 Tab",                          "ADMIN", DONE, ""),
    ("发现页配置", "GET",    "/discovery/settings",               "全局文案 K-V",                      "公开", DONE, ""),
    ("发现页配置", "PUT",    "/discovery/settings/{key}",         "Upsert 全局文案",                   "ADMIN", DONE, ""),
    # 字典
    ("字典", "GET",    "/dict",                                   "字典列表",                          "需登录", DONE, ""),
    ("字典", "POST",   "/dict",                                   "新建字典项",                        "ADMIN", DONE, ""),
    ("字典", "PUT",    "/dict/{id}",                              "更新字典项",                        "ADMIN", DONE, ""),
    ("字典", "DELETE", "/dict/{id}",                              "删除字典项",                        "ADMIN", DONE, ""),
    ("字典", "POST",   "/dict/reload",                            "刷新字典缓存",                      "ADMIN", DONE, ""),
    # 反馈
    ("反馈工单", "GET",    "/feedback/tickets",                   "我的工单",                          "需登录", DONE, ""),
    ("反馈工单", "GET",    "/feedback/tickets/{id}",              "工单详情",                          "需登录", DONE, ""),
    ("反馈工单", "POST",   "/feedback/tickets",                   "创建工单",                          "需登录", DONE, ""),
    ("反馈工单", "POST",   "/feedback/tickets/{id}/reply",        "用户追加回复",                      "需登录", DONE, ""),
]

for i, (mod, m, path, desc, auth, status, note) in enumerate(endpoints, 2):
    write_row(ws2, i, [i-1, mod, m, path, desc, auth, status, note], 7)

# ============================================================
# Sheet 3: 前端页面
# ============================================================
ws3 = wb.create_sheet("前端页面")
write_header(ws3, ["#", "分组", "页面", "说明", "状态", "备注"],
             [4, 12, 26, 36, 12, 30])

screens = [
    ("发现",   "HomeScreen",            "瀑布流 + 搜索 + 分类筛选 + 排序",         DONE, ""),
    ("创作",   "CreateScreen",          "创作入口（手动/AI/图片）",                 DONE, ""),
    ("创作",   "EditorScreen",          "Canvas 拼豆编辑器",                         DONE, ""),
    ("市场",   "MarketScreen",          "材料商城 + 图纸市场 tab 切换",              DONE, ""),
    ("市场",   "ProductDetailScreen",   "商品详情",                                  DONE, ""),
    ("市场",   "CartScreen",            "购物车",                                    DONE, ""),
    ("市场",   "PaymentScreen",         "结算 / 支付页",                             PART, "UI 完成；真实支付待接"),
    ("详情",   "DesignDetailScreen",    "作品详情",                                  DONE, ""),
    ("详情",   "ResourceDetailScreen",  "资源详情（图纸 / 商品共用）",                DONE, ""),
    ("我的",   "ProfileScreen",         "个人中心",                                  DONE, ""),
    ("我的",   "LoginScreen",           "登录",                                      DONE, ""),
    ("我的",   "RegisterScreen",        "注册",                                      DONE, ""),
    ("我的",   "EditProfileScreen",     "编辑资料 + 头像上传",                       DONE, ""),
    ("我的",   "MyDesignsScreen",       "我的作品",                                  DONE, ""),
    ("我的",   "MyFeedsScreen",         "我的动态",                                  DONE, ""),
    ("我的",   "FavoritesScreen",       "我的收藏",                                  DONE, ""),
    ("我的",   "LikedHistoryScreen",    "我点过的赞",                                DONE, ""),
    ("我的",   "LikesScreen",           "我收到的赞",                                DONE, ""),
    ("我的",   "WalletScreen",          "钱包 + 签到 + 流水",                        DONE, ""),
    ("我的",   "OrdersScreen",          "订单列表（含状态过滤）",                    DONE, ""),
    ("我的",   "OrderDetailScreen",     "订单详情",                                  DONE, ""),
    ("我的",   "PurchasedScreen",       "已购图纸",                                  DONE, ""),
    ("我的",   "AddressScreen",         "地址列表（结算用）",                        DONE, ""),
    ("我的",   "AddressManageScreen",   "地址管理 + 默认设置",                       DONE, ""),
    ("我的",   "FollowListScreen",      "粉丝/关注列表",                             DONE, ""),
    ("我的",   "NotificationsScreen",   "通知中心",                                  DONE, ""),
    ("我的",   "FeedbackScreen",        "反馈工单列表",                              DONE, ""),
    ("我的",   "FeedbackDetailScreen",  "工单详情/回复",                             DONE, ""),
    ("我的",   "SettingsScreen",        "设置（改密 / 退出 / 隐私）",                DONE, ""),
    ("社区",   "DirectMessageScreen",   "私信会话页",                                TODO, "无后端 DM 模块；纯 UI"),
]

for i, (group, screen, desc, status, note) in enumerate(screens, 2):
    write_row(ws3, i, [i-1, group, screen, desc, status, note], 5)

# ============================================================
# Sheet 4: 管理后台
# ============================================================
ws4 = wb.create_sheet("管理后台")
write_header(ws4, ["#", "页面", "功能", "状态", "备注"], [4, 18, 36, 12, 32])

admin_pages = [
    ("Dashboard",  "全局统计图表（用户/作品/订单等）",          DONE, "ECharts"),
    ("Users",      "用户列表 / 搜索 / 删除",                    DONE, ""),
    ("Designs",    "作品列表 / 删除",                           DONE, ""),
    ("Products",   "材料商品 CRUD",                             DONE, ""),
    ("Patterns",   "图纸下架（删除）",                          PART, "缺新增/编辑/上下架"),
    ("Feeds",      "动态删除",                                  PART, "缺编辑/置顶/精选"),
    ("ApiConfig",  "豆包 / 微信 / 支付宝密钥配置（脱敏）",      DONE, "支持 reveal + 审计日志"),
    ("Discovery",  "发现页 Banner + Tab 拖拽排序",              DONE, "近期新增"),
    ("Dict",       "字典管理（订单状态 / 工单类型 等）",         DONE, "近期新增；写后自动 reload"),
    ("Login",      "管理员登录",                                DONE, ""),
    ("Feedback",   "反馈工单管理 + 客服回复",                   TODO, "前端无此页面"),
    ("Orders",     "订单运营（发货 / 退款审核）",                TODO, "缺管理员入口"),
]

for i, (page, fn, status, note) in enumerate(admin_pages, 2):
    write_row(ws4, i, [i-1, page, fn, status, note], 4)

# ============================================================
# Sheet 5: 待办 / 缺口
# ============================================================
ws5 = wb.create_sheet("待办与缺口")
write_header(ws5, ["#", "事项", "影响范围", "优先级", "备注"], [4, 36, 18, 10, 50])

todos = [
    ("微信 / 支付宝真实支付接入",         "充值 / 订单付款",  "高", "PaymentController 有 TODO 占位；需要下单 API + 回调验签 + 防重放"),
    ("订单退款链路真实落地",              "订单",             "中", "目前 /orders/{id}/refund 只改状态，不退钱"),
    ("私信 (DM) 后端模块",                "社区",             "中", "前端 DirectMessageScreen 已存在；需建 Conversation / Message 表"),
    ("注册邮箱验证",                      "认证",             "中", "register 当前不校验邮箱，可能导致脏注册"),
    ("管理后台 — 反馈工单回复页",         "运营",             "中", "用户能创建工单但 admin 没办法回复"),
    ("管理后台 — 订单运营页",             "运营",             "中", "缺发货 / 退款审核入口"),
    ("管理后台 — 图纸完整管理",           "运营",             "低", "目前只能删除"),
    ("管理后台 — 动态完整管理",           "运营",             "低", "目前只能删除"),
    ("反馈工单截图字段持久化",            "反馈",             "低", "screenshots 字段已写入但未在 DTO 显示"),
    ("/uploads 文件清理 / OSS 化",        "上传",             "低", "本地落盘，重启容器丢失；建议 MinIO/OSS"),
    ("CI 增加单元测试",                    "工程化",           "低", "目前只 build Docker，没跑测试"),
]
for i, (item, scope, pri, note) in enumerate(todos, 2):
    pri_cell = pri
    write_row(ws5, i, [i-1, item, scope, pri_cell, note], 4)
    # 优先级染色
    fill = PatternFill("solid", fgColor="F1948A") if pri == "高" else (
           PatternFill("solid", fgColor="F8C471") if pri == "中" else
           PatternFill("solid", fgColor="A9DFBF"))
    ws5.cell(row=i, column=4).fill = fill

# ---------- 保存 ----------
wb.save(OUT)
print(f"[OK] -> {OUT}")
